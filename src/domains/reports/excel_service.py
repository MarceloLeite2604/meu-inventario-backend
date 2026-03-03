from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..inventories.models import Inventario
from ..organizations.models import Organizacao
from ..scope1.effluents.models import EmissaoEfluente
from ..scope1.fugitive.models import EmissaoFugitiva
from ..scope1.mobile_combustion.models import EmissaoCombustaoMovel
from ..scope1.stationary_combustion.models import EmissaoEstacionaria
from ..scope2.energy.models import ConsumoEnergia
from ..scope3.business_travel.models import EmissaoViagemNegocio

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SCOPE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_SCOPE_FONT = Font(color="FFFFFF", bold=True)
_TOTAL_FONT = Font(bold=True)


def _header(ws, row: int, cols: list[str]) -> None:
    for col, title in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _scope_row(ws, row: int, label: str, n_cols: int) -> None:
    ws.cell(row=row, column=1, value=label).font = _SCOPE_FONT
    ws.cell(row=row, column=1).fill = _SCOPE_FILL
    for col in range(2, n_cols + 1):
        ws.cell(row=row, column=col).fill = _SCOPE_FILL


async def generate_excel(inventory_id: UUID, session: AsyncSession) -> bytes:
    inv_result = await session.execute(
        select(Inventario).where(Inventario.id == inventory_id))
    inv = inv_result.scalar_one_or_none()

    org_name = ""
    ano_base = ""
    org_id = None
    if inv:
        org_id = inv.organizacao_id
        ano_base = str(inv.ano_base)
        org_result = await session.execute(
            select(Organizacao).where(Organizacao.id == inv.organizacao_id))
        org = org_result.scalar_one_or_none()
        org_name = org.nome if org else ""

    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Resumo"

    ws_summary["A1"] = "Inventario de Emissoes GHG"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = f"Organizacao: {org_name}"
    ws_summary["A3"] = f"Ano base: {ano_base}"
    ws_summary.append([])

    _header(ws_summary, 5, ["Escopo", "Descricao", "Emissoes (tCO2e)"])

    scope_rows: list[tuple[str, str, float]] = []

    if org_id:
        async def org_sum(model, col):
            result = await session.execute(
                select(model).where(model.organizacao_id == org_id))
            return sum((getattr(r, col) or 0.0) for r in result.scalars().all())

        mobile = await org_sum(EmissaoCombustaoMovel, "emissoes_tco2e_total")
        stationary = await org_sum(EmissaoEstacionaria, "emissoes_tco2e_total")
        fugitive = await org_sum(EmissaoFugitiva, "emissoes_tco2e")
        effluent = await org_sum(EmissaoEfluente, "emissoes_tco2e_total")
        energy = await org_sum(ConsumoEnergia, "emissoes_energia_tco2e")
        travel = await org_sum(EmissaoViagemNegocio, "emissoes_tco2e_total")

        scope_rows = [
            ("Escopo 1", "Combustao Movel", mobile),
            ("Escopo 1", "Combustao Estacionaria", stationary),
            ("Escopo 1", "Emissoes Fugitivas", fugitive),
            ("Escopo 1", "Efluentes", effluent),
            ("Escopo 2", "Energia Eletrica", energy),
            ("Escopo 3", "Viagens de Negocios", travel),
        ]

    row = 6
    for scope, desc, value in scope_rows:
        ws_summary.cell(row=row, column=1, value=scope)
        ws_summary.cell(row=row, column=2, value=desc)
        ws_summary.cell(row=row, column=3, value=round(value, 4))
        row += 1

    total = sum(v for _, _, v in scope_rows)
    total_row = row
    ws_summary.cell(total_row, 1, "TOTAL").font = _TOTAL_FONT
    ws_summary.cell(total_row, 3, round(total, 4)).font = _TOTAL_FONT

    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 35
    ws_summary.column_dimensions["C"].width = 20

    # ── Scope 1 detail sheet ───────────────────────────────────────────────
    ws1 = wb.create_sheet("Escopo 1")

    if org_id:
        # Mobile combustion
        result = await session.execute(
            select(EmissaoCombustaoMovel).where(EmissaoCombustaoMovel.organizacao_id == org_id))
        mobile_records = result.scalars().all()

        _scope_row(ws1, 1, "Combustao Movel", 6)
        _header(ws1, 2, ["Ano", "Mes", "Combustivel", "Quantidade", "Unidade", "Total tCO2e"])
        for row_num, r in enumerate(mobile_records, start=3):
            ws1.cell(row_num, 1, r.ano)
            ws1.cell(row_num, 2, r.mes)
            ws1.cell(row_num, 3, r.combustivel)
            ws1.cell(row_num, 4, r.quantidade_combustivel)
            ws1.cell(row_num, 5, r.unidade_combustivel)
            ws1.cell(row_num, 6, r.emissoes_tco2e_total)

    # ── Scope 2 detail sheet ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Escopo 2")

    if org_id:
        result = await session.execute(
            select(ConsumoEnergia).where(ConsumoEnergia.organizacao_id == org_id))
        energy_records = result.scalars().all()

        _scope_row(ws2, 1, "Consumo de Energia Eletrica", 5)
        _header(ws2, 2, ["Ano", "Mes", "Consumo (MWh)", "Emissoes (tCO2e)", "Descricao"])
        for row_num, r in enumerate(energy_records, start=3):
            ws2.cell(row_num, 1, r.ano)
            ws2.cell(row_num, 2, r.mes)
            ws2.cell(row_num, 3, r.consumo_mwh)
            ws2.cell(row_num, 4, r.emissoes_energia_tco2e)
            ws2.cell(row_num, 5, r.descricao)

    # ── Scope 3 detail sheet ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Escopo 3")

    if org_id:
        result = await session.execute(
            select(EmissaoViagemNegocio).where(EmissaoViagemNegocio.organizacao_id == org_id))
        travel_records = result.scalars().all()

        _scope_row(ws3, 1, "Viagens de Negocios", 7)
        _header(ws3, 2, ["Ano", "Mes", "Tipo Transporte", "Origem", "Destino", "Distancia (km)", "Total tCO2e"])
        for row_num, r in enumerate(travel_records, start=3):
            ws3.cell(row_num, 1, r.ano)
            ws3.cell(row_num, 2, r.mes)
            ws3.cell(row_num, 3, r.tipo_transporte)
            ws3.cell(row_num, 4, r.origin)
            ws3.cell(row_num, 5, r.destination)
            ws3.cell(row_num, 6, r.distance)
            ws3.cell(row_num, 7, r.emissoes_tco2e_total)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
