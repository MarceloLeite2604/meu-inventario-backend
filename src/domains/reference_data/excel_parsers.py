import unicodedata
from io import BytesIO

from openpyxl import load_workbook

_MONTH_NAMES: dict[str, int] = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Normalized sheet name → target table name
_SHEET_TO_TABLE: dict[str, str] = {
    "composicao_combustiveis": "composicao_combustiveis",
    "equivalencia_veiculos": "equivalencia_veiculos",
    "todos_combustiveis": "fatores_frota_tipo_combustivel",
    "fatores_frota_tipo_combustivel": "fatores_frota_tipo_combustivel",
    "gwp": "gwp",
    "fatores_variaveis_ghg": "fatores_variaveis_ghg",
    "variaveis_ghg": "fatores_variaveis_ghg",
    "transporte_metro": "transporte_metro",
    "transporte_trem": "transporte_trem",
    "consumo_unidade_medida": "consumo_unidade_medida",
    "fatores_estacionaria": "fatores_estacionaria",
    "fatores_tipo_combustivel": "fatores_tipo_combustivel",
    "fatores_emissao_aereas": "fatores_emissao_aereas",
    "fatores_transporte_onibus": "fatores_transporte_onibus",
    "transporte_onibus": "fatores_transporte_onibus",
}


def normalize_col(name: str) -> str:
    """Lowercase, remove accents, replace spaces/hyphens with underscores."""
    nfkd = unicodedata.normalize("NFKD", str(name))
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    return ascii_str.lower().strip().replace(" ", "_").replace("-", "_")


def _to_float(val: object) -> float | None:
    if val is None:
        return None
    try:
        return float(val)  # type: ignore[arg-type]
    except (ValueError, TypeError):
        return None


def _dms_to_decimal(graus: float, minutos: float, segundos: float, direcao: str) -> float:
    decimal = graus + minutos / 60 + segundos / 3600
    if direcao.upper() in ("S", "O", "W"):
        decimal = -decimal
    return round(decimal, 6)


def parse_energy_factors(content: bytes) -> list[dict]:
    """Parse ANEEL-format energy emission factor spreadsheet.

    For each integer year found in the first cell of a row, searches the
    next 3 rows for month column headers (Janeiro–Dezembro) and reads the
    values row immediately after.
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records: list[dict] = []
    i = 0
    while i < len(rows):
        row = rows[i]
        first = row[0] if row else None
        if isinstance(first, int) and 1900 < first < 2200:
            ano = first
            header_row_idx: int | None = None
            for j in range(i + 1, min(i + 4, len(rows))):
                candidate = rows[j]
                month_found = any(
                    normalize_col(str(c)) in _MONTH_NAMES
                    for c in candidate
                    if c is not None
                )
                if month_found:
                    header_row_idx = j
                    break
            if header_row_idx is not None and header_row_idx + 1 < len(rows):
                header = rows[header_row_idx]
                values_row = rows[header_row_idx + 1]
                for col_idx, cell_name in enumerate(header):
                    if cell_name is None:
                        continue
                    mes = _MONTH_NAMES.get(normalize_col(str(cell_name)))
                    if mes is None:
                        continue
                    val = values_row[col_idx] if col_idx < len(values_row) else None
                    fator = _to_float(val)
                    if fator is None:
                        continue
                    records.append({"ano": ano, "mes": mes, "fator_emissao": fator})
                i = header_row_idx + 2
                continue
        i += 1
    return records


def parse_airports(content: bytes) -> dict[str, list[dict]]:
    """Parse airport Excel: 'Aeroportos' sheet (DMS coords) and optional aerial factors sheet."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    result: dict[str, list[dict]] = {"aeroportos": [], "fatores_emissao_aereas": []}

    for sheet_name in wb.sheetnames:
        normalized_sheet = normalize_col(sheet_name)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header = [normalize_col(str(c)) if c is not None else "" for c in rows[0]]

        if normalized_sheet == "aeroportos":
            for row in rows[1:]:
                if not any(c is not None for c in row):
                    continue
                d = {k: v for k, v in zip(header, row) if k}
                sigla = d.get("sigla")
                if not sigla:
                    continue
                gl = _to_float(d.get("graus_lat"))
                ml = _to_float(d.get("minutos_lat")) or 0.0
                sl = _to_float(d.get("segundos_lat")) or 0.0
                dl = str(d.get("direcao_lat") or "N").strip()
                glo = _to_float(d.get("graus_lon"))
                mlo = _to_float(d.get("minutos_lon")) or 0.0
                slo = _to_float(d.get("segundos_lon")) or 0.0
                dlo = str(d.get("direcao_lon") or "W").strip()
                if gl is None or glo is None:
                    continue
                result["aeroportos"].append({
                    "sigla": str(sigla).strip().upper(),
                    "nome": d.get("nome"),
                    "latitude": _dms_to_decimal(gl, ml, sl, dl),
                    "longitude": _dms_to_decimal(glo, mlo, slo, dlo),
                    "graus_lat": gl,
                    "minutos_lat": ml,
                    "segundos_lat": sl,
                    "direcao_lat": dl[:1],
                    "graus_lon": glo,
                    "minutos_lon": mlo,
                    "segundos_lon": slo,
                    "direcao_lon": dlo[:1],
                })

        elif normalized_sheet in ("fatores_emissao_aereas", "fatores_emissao_aerea"):
            for row in rows[1:]:
                if not any(c is not None for c in row):
                    continue
                d = {k: v for k, v in zip(header, row) if k}
                ano = d.get("ano_referencia")
                distancia = d.get("distancia_aerea")
                if ano is None or distancia is None:
                    continue
                result["fatores_emissao_aereas"].append({
                    "ano_referencia": int(ano),
                    "distancia_aerea": str(distancia),
                    "acrescimo_rota": _to_float(d.get("acrescimo_rota")) or 0.0,
                    "co2_aereo_passageiro_km": _to_float(d.get("co2_aereo_passageiro_km")) or 0.0,
                    "ch4_aereo_passageiro_km": _to_float(d.get("ch4_aereo_passageiro_km")) or 0.0,
                    "n2o_aereo_passageiro_km": _to_float(d.get("n2o_aereo_passageiro_km")) or 0.0,
                })

    wb.close()
    return result


def parse_vehicle_factors(content: bytes) -> dict[str, list[dict]]:
    """Parse multi-sheet vehicle factors Excel into table_name → list of row dicts."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        table_name = _SHEET_TO_TABLE.get(normalize_col(sheet_name))
        if not table_name:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header = [normalize_col(str(c)) if c is not None else "" for c in rows[0]]
        table_rows: list[dict] = []
        for row in rows[1:]:
            if not any(c is not None for c in row):
                continue
            table_rows.append({k: v for k, v in zip(header, row) if k})
        if table_rows:
            if table_name in result:
                result[table_name].extend(table_rows)
            else:
                result[table_name] = table_rows

    wb.close()
    return result
