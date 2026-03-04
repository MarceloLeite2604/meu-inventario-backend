"""Tests for bulk reference data import (energy factors, airports, vehicle factors)."""

import io
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from src.domains.reference_data import service
from src.domains.reference_data.excel_parsers import (
    normalize_col,
    parse_airports,
    parse_energy_factors,
    parse_vehicle_factors,
)
from src.domains.reference_data.models import (
    AeroportoCoordenada,
    FatorEmissaoEnergia,
    Gwp,
)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_energy_factors_excel() -> bytes:
    """Create an ANEEL-style energy factors Excel with 2 years."""
    wb = Workbook()
    ws = wb.active
    # Year 2022
    ws.append([2022])
    ws.append(["Legenda", "Unidade"])
    ws.append(
        [None, None, "Janeiro", "Fevereiro", "Março", "Abril",
         "Maio", "Junho", "Julho", "Agosto",
         "Setembro", "Outubro", "Novembro", "Dezembro"]
    )
    ws.append(
        [None, None, 0.1, 0.2, 0.3, 0.4,
         0.5, 0.6, 0.7, 0.8,
         0.9, 1.0, 1.1, 1.2]
    )
    # Year 2023
    ws.append([2023])
    ws.append(["Legenda", "Unidade"])
    ws.append(
        [None, None, "Janeiro", "Fevereiro", "Março"]
    )
    ws.append([None, None, 0.15, 0.25, 0.35])
    return _wb_to_bytes(wb)


def _make_airport_excel() -> bytes:
    """Create an airport Excel with Aeroportos sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Aeroportos"
    ws.append(["sigla", "nome", "graus_lat", "minutos_lat", "segundos_lat", "direcao_lat",
               "graus_lon", "minutos_lon", "segundos_lon", "direcao_lon"])
    ws.append(["GRU", "Guarulhos", 23, 25, 50, "S", 46, 28, 22, "O"])
    ws.append(["CGH", "Congonhas", 23, 37, 38, "S", 46, 39, 21, "O"])
    return _wb_to_bytes(wb)


def _make_vehicle_factors_excel() -> bytes:
    """Create a vehicle factors Excel with gwp sheet."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    ws_gwp = wb.create_sheet("gwp")
    ws_gwp.append(["nome_ghg", "gwp_value"])
    ws_gwp.append(["CO2", 1.0])
    ws_gwp.append(["CH4", 28.0])
    ws_gwp.append(["N2O", 265.0])
    ws_metro = wb.create_sheet("transporte_metro")
    ws_metro.append(["ano", "g_co2_passageiro_km"])
    ws_metro.append([2022, 4.5])
    ws_metro.append([2023, 4.2])
    return _wb_to_bytes(wb)


# ─── Parser unit tests ──────────────────────────────────────────────────────

class TestNormalizeCol:
    def test_removes_accents(self):
        assert normalize_col("Março") == "marco"

    def test_lowercases(self):
        assert normalize_col("Janeiro") == "janeiro"

    def test_replaces_spaces_with_underscores(self):
        assert normalize_col("graus lat") == "graus_lat"

    def test_replaces_hyphens_with_underscores(self):
        assert normalize_col("co2-fator") == "co2_fator"


class TestParseEnergyFactors:
    def test_parses_year_and_months(self):
        content = _make_energy_factors_excel()
        records = parse_energy_factors(content)

        anos = {r["ano"] for r in records}
        assert 2022 in anos
        assert 2023 in anos

    def test_returns_correct_count(self):
        content = _make_energy_factors_excel()
        records = parse_energy_factors(content)

        # 12 months from 2022 + 3 months from 2023
        assert len(records) == 15

    def test_record_structure(self):
        content = _make_energy_factors_excel()
        records = parse_energy_factors(content)

        jan_2022 = next(r for r in records if r["ano"] == 2022 and r["mes"] == 1)
        assert jan_2022["fator_emissao"] == pytest.approx(0.1)

    def test_empty_bytes_returns_empty(self):
        wb = Workbook()
        wb.active.append(["just text"])
        content = _wb_to_bytes(wb)
        records = parse_energy_factors(content)
        assert records == []


class TestParseAirports:
    def test_parses_aeroportos_sheet(self):
        content = _make_airport_excel()
        result = parse_airports(content)

        assert len(result["aeroportos"]) == 2

    def test_converts_dms_to_decimal(self):
        content = _make_airport_excel()
        result = parse_airports(content)

        gru = next(a for a in result["aeroportos"] if a["sigla"] == "GRU")
        # 23° 25' 50" S → negative
        assert gru["latitude"] < 0
        assert abs(gru["latitude"]) == pytest.approx(23 + 25 / 60 + 50 / 3600, rel=1e-4)

    def test_no_aerial_factors_sheet_returns_empty_list(self):
        content = _make_airport_excel()
        result = parse_airports(content)

        assert result["fatores_emissao_aereas"] == []

    def test_airport_sigla_uppercased(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Aeroportos"
        ws.append(["sigla", "nome", "graus_lat", "minutos_lat", "segundos_lat",
                   "direcao_lat", "graus_lon", "minutos_lon", "segundos_lon", "direcao_lon"])
        ws.append(["gru", "Guarulhos", 23, 25, 50, "S", 46, 28, 22, "O"])
        content = _wb_to_bytes(wb)
        result = parse_airports(content)
        assert result["aeroportos"][0]["sigla"] == "GRU"


class TestParseVehicleFactors:
    def test_parses_known_sheet(self):
        content = _make_vehicle_factors_excel()
        result = parse_vehicle_factors(content)

        assert "gwp" in result
        assert len(result["gwp"]) == 3

    def test_parses_multiple_sheets(self):
        content = _make_vehicle_factors_excel()
        result = parse_vehicle_factors(content)

        assert "transporte_metro" in result
        assert len(result["transporte_metro"]) == 2

    def test_unknown_sheet_ignored(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "unknown_sheet_xyz"
        ws.append(["col1", "col2"])
        ws.append([1, 2])
        content = _wb_to_bytes(wb)
        result = parse_vehicle_factors(content)
        assert result == {}


# ─── Service integration tests ──────────────────────────────────────────────

@pytest.mark.anyio
class TestImportEnergyFactorsService:
    async def test_imports_and_replaces(self, database_session: AsyncSession):
        # Insert an existing record that should be replaced
        database_session.add(FatorEmissaoEnergia(ano=2020, mes=1, fator_emissao=99.0))
        await database_session.flush()

        content = _make_energy_factors_excel()
        count = await service.import_energy_factors(content, database_session)

        assert count == 15
        # Old record should be gone; verify by checking the new ones
        from sqlalchemy import select
        result = await database_session.execute(select(FatorEmissaoEnergia))
        rows = list(result.scalars().all())
        anos = {r.ano for r in rows}
        assert 2020 not in anos
        assert 2022 in anos

    async def test_returns_correct_count(self, database_session: AsyncSession):
        content = _make_energy_factors_excel()
        count = await service.import_energy_factors(content, database_session)
        assert count == 15


@pytest.mark.anyio
class TestImportAirportsService:
    async def test_imports_airports(self, database_session: AsyncSession):
        content = _make_airport_excel()
        counts = await service.import_airports(content, database_session)

        assert counts["aeroportos"] == 2
        assert counts["fatores_emissao_aereas"] == 0

    async def test_replaces_existing_airports(self, database_session: AsyncSession):
        database_session.add(AeroportoCoordenada(
            sigla="OLD", latitude=-10.0, longitude=-50.0
        ))
        await database_session.flush()

        content = _make_airport_excel()
        await service.import_airports(content, database_session)

        from sqlalchemy import select
        result = await database_session.execute(select(AeroportoCoordenada))
        rows = list(result.scalars().all())
        siglas = {r.sigla for r in rows}
        assert "OLD" not in siglas
        assert "GRU" in siglas


@pytest.mark.anyio
class TestImportVehicleFactorsService:
    async def test_imports_gwp_rows(self, database_session: AsyncSession):
        content = _make_vehicle_factors_excel()
        counts = await service.import_vehicle_factors(content, database_session)

        assert counts.get("gwp") == 3

    async def test_replaces_gwp_rows(self, database_session: AsyncSession):
        database_session.add(Gwp(nome_ghg="OLD_GAS", gwp_value=999.0))
        await database_session.flush()

        content = _make_vehicle_factors_excel()
        await service.import_vehicle_factors(content, database_session)

        from sqlalchemy import select
        result = await database_session.execute(select(Gwp))
        rows = list(result.scalars().all())
        names = {r.nome_ghg for r in rows}
        assert "OLD_GAS" not in names
        assert "CH4" in names


# ─── Route integration tests ─────────────────────────────────────────────────

INVENTORY_ID = str(uuid4())


@pytest.mark.anyio
class TestImportEnergyFactorsRoute:
    async def test_import_returns_200(self, client: AsyncClient):
        content = _make_energy_factors_excel()
        response = await client.post(
            "/reference-data/energy-factors/import",
            files={"file": ("fatores.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["importados"] == 15

    async def test_import_empty_excel_returns_zero(self, client: AsyncClient):
        wb = Workbook()
        wb.active.append(["no year rows"])
        content = _wb_to_bytes(wb)
        response = await client.post(
            "/reference-data/energy-factors/import",
            files={"file": ("fatores.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["importados"] == 0


@pytest.mark.anyio
class TestImportAirportsRoute:
    async def test_import_returns_200(self, client: AsyncClient):
        content = _make_airport_excel()
        response = await client.post(
            "/reference-data/airports/import",
            files={"file": ("aeroportos.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["aeroportos"] == 2
        assert data["fatores_emissao_aereas"] == 0


@pytest.mark.anyio
class TestImportVehicleFactorsRoute:
    async def test_import_returns_200(self, client: AsyncClient):
        content = _make_vehicle_factors_excel()
        response = await client.post(
            "/reference-data/vehicle-factors/import",
            files={"file": ("fatores.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert "gwp" in data["tabelas"]
        assert data["tabelas"]["gwp"] == 3
